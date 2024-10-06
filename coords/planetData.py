import openpyxl
import numpy as np


class PlanetData:
    name = []
    year = []
    ra = []
    dec = []
    dist = []
    x = []
    y = []
    z = []

    def __init__(self, file):
        dataFrame = openpyxl.load_workbook(file).active

        for row in range(1, dataFrame.max_row):
            for idx, col in enumerate(dataFrame.iter_cols(1, dataFrame.max_column)):
                if idx == 0:
                    self.name.append(col[row].value)
                elif idx == 1:
                    self.year.append(col[row].value)
                elif idx == 2:
                    self.ra.append(col[row].value)
                elif idx == 3:
                    self.dec.append(col[row].value)
                elif idx == 4:
                    self.dist.append(col[row].value)


def TripleDtoDoubleD(x, y, z, StarX, StarY, StarZ):
    # arrays of star and exoplanet position
    positionStar = np.array([StarX, StarY, StarZ])
    cameraPosition = np.array([x, y, z])

    # unitvector point straight ahead from origin (camera position/exoplanet position)
    # and its magnitude
    unitvecY = np.array([x, y + 1, z])
    unitvecYMag = np.linalg.norm(unitvecY)

    # field of views in vertical and horizontal direction
    FOVz = 180 / 2 * (np.pi / 180)
    FOVx = 180 / 2 * (np.pi / 180)

    # width and height of screen (set to 1920x1080)
    widthX = 1920 / 2
    heightY = 1080 / 2

    # vector from exoplanet to star & magnitude of the vector
    distanceVec = np.subtract(positionStar, cameraPosition)
    distanceMag = np.linalg.norm(distanceVec)

    dVecX = distanceVec.copy()
    dVecZ = distanceVec.copy()

    # vectors in x-y plane to find angle with unit vector
    dVecX[2] = z
    dVecZ[0] = x

    # take the magnitudes of the vectors
    dVecXMag = np.linalg.norm(dVecX)
    dVecZMag = np.linalg.norm(dVecZ)

    # Find the horizontal angle
    inside1 = np.dot(unitvecY, dVecX) / (dVecXMag * unitvecYMag)
    angleX = np.arccos(inside1)

    # Find the vertical angle
    inside2 = np.dot(unitvecY, dVecZ) / (dVecZMag * unitvecYMag)
    angleZ = np.arccos(inside2)

    # Find the coords on the screen based on the angle fraction
    fracX = angleX / FOVx
    fracZ = angleZ / FOVz

    # coords on the plane!
    # xCoord = fracX * widthX +widthX/2
    # yCoord = (fracZ * heightY)

    # coords on the plane!
    xCoord = fracX * widthX
    if StarX < x:
        xCoord = widthX - xCoord
    else:
        xCoord += widthX

    yCoord = fracZ * heightY
    if StarZ < z:
        yCoord += heightY
    else:
        yCoord = heightY - yCoord

    xint = int(xCoord)
    yint = int(yCoord)

    return xint, yint